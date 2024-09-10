import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import Cell
from collections import defaultdict
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# Style definitions
title_font = Font(size=14, bold=True)
title_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
border_style = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)


bold_font = Font(bold=True)
centered_alignment = Alignment(horizontal='center', vertical='center')

def create_table(ws, start_row, end_row, start_col, end_col, table_name):
    """Create a table in the worksheet for the specified range with a unique name."""
    table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    
    # Define the table with a unique name
    table = Table(displayName=table_name, ref=table_range)
    
    # Apply a medium style with banded rows
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    
    # Add the table to the worksheet
    ws.add_table(table)


def extract_heading_from_filepath(file_path):
    """Extract the heading from the file name in the file path."""
    file_name = os.path.basename(file_path)
    heading, _ = os.path.splitext(file_name)
    return heading

def update_dataframe(file_path, df,base_value,module_name,addr_val):

  # Insert new columns b31 to b0
    new_column_names = [f'b{i}' for i in range(31, -1, -1)]
    for i, column_name in enumerate(new_column_names):
        df.insert(3+i, column_name, None)

    # Modify 'Address Offset' column
    modify_column = 'Address Offset'
    if modify_column in df.columns:
        df[modify_column] = df[modify_column].str.replace('0x', '', regex=False)
        df.rename(columns={'Address Offset': 'FPGA Addr offset (in HEX)'}, inplace=True)

    # Rename columns if they exist
    if 'Register Description' in df.columns:
        df.rename(columns={'Register Description': 'Notes'}, inplace=True)
    if 'Access' in df.columns:
        df.rename(columns={'Access': 'Permission'}, inplace=True)

    # Drop rows with NaN in the first column
    df = df.dropna(subset=[df.columns[1]])  # Adjusted to drop based on the actual first data column

    
    df.iloc.insert(2, 'Register Size in bytes', 4)
    df.iloc.insert(len(df.columns), 'Base Address', addr_val)
    df.iloc.insert(len(df.columns), 'Subsystem Bridge', ' ')
    df.iloc.insert(len(df.columns), 'Top Qsys HPS Bridge', 'hps_h2f_lw_axi_master')
    df.iloc.insert(len(df.columns), 'Actual Base Address = Top Qsys HPS Bridge Base Address + Subsystem Base Address',base_value)
  
    if module_name=="DFD Subsystem Registers":
     df['Top Qsys HPS Bridge'] ="0x0004_4000"
    if module_name=="TOD Timestamp Buffer - Registers":
        df['Top Qsys HPS Bridge'] ="0X0005_0000"
    else:    
         df['Top Qsys HPS Bridge'] ="0x0010_0000"
    return df

def get_merged_cell_ranges(ws):
    """Gets merged cell ranges and maps each cell to its merged value."""
    merged_ranges = ws.merged_cells.ranges
    merged_cells = {}
    for merged_range in merged_ranges:
        start_cell = merged_range.start_cell
        module_name = ws[start_cell.coordinate].value
        for row in ws.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row, min_col=merged_range.min_col, max_col=merged_range.max_col):
            for cell in row:
                merged_cells[cell.coordinate] = module_name
    return merged_cells

def find_first_and_last_position_per_module(file_path):
    """Finds the first and last row index for each module across all sheets."""
    wb = load_workbook(file_path)
    sheet_results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged_cells = get_merged_cell_ranges(ws)
        module_positions = {}

        for row in ws.iter_rows(min_col=2, max_col=2):  # Assuming 'Module Name' is in Column B
            cell = row[0]
            if cell.row > 1:  # Skip header
                module_name = merged_cells.get(cell.coordinate, cell.value)
                if module_name:
                    if module_name not in module_positions:
                        module_positions[module_name] = {'first': cell.row, 'last': cell.row}
                    else:
                        module_positions[module_name]['last'] = cell.row

        sheet_results[sheet_name] = module_positions
    return sheet_results

def insert_module_title(ws, module_name, start_row, start_col, end_col, title_style):
   
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = module_name
    title_cell.font = Font(size=14, bold=True)
    title_cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    # Apply the alignment to top-left
    title_cell.alignment = Alignment(horizontal='left', vertical='top')
    # Apply the additional alignment style passed in the function arguments
    title_cell.alignment = title_style
    
def merge_adjacent_empty_cells(ws, start_row, end_row, start_col, end_col):
    """Merges adjacent empty cells in each row."""
    for row in range(start_row, end_row + 1):
        merge_start_col = None
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value in [None, ""]:
                if merge_start_col is None:
                    merge_start_col = col
            else:
                if merge_start_col is not None:
                    ws.merge_cells(start_row=row, start_column=merge_start_col, end_row=row, end_column=col - 1)
                    merge_start_col = None
        if merge_start_col is not None:
            ws.merge_cells(start_row=row, start_column=merge_start_col, end_row=row, end_column=end_col)

def apply_column_merges(ws, start_row, start_col, end_row, end_col, columns_to_merge):
    """Merges specific columns across the given row range."""
    for col in columns_to_merge:
        col_index = column_index_from_string(col)
        ws.merge_cells(start_row=start_row, start_column=col_index, end_row=end_row, end_column=col_index)
        merged_cell = ws.cell(row=start_row, column=col_index)
        merged_cell.value = ws.cell(row=start_row, column=col_index).value

def apply_border_and_alignment(ws, start_row, end_row, start_col, end_col, border_style, alignment_style):
    """Applies border and alignment to the specified cell range."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border_style
            cell.alignment = alignment_style

def merge_adjacent_empty_cells(ws, start_row, end_row, start_col, end_col):
    """Merges adjacent empty cells in each row."""
    for row in range(start_row, end_row + 1):
        merge_start_col = None
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value in [None, ""]:
                if merge_start_col is None:
                    merge_start_col = col
            else:
                if merge_start_col is not None:
                    ws.merge_cells(start_row=row, start_column=merge_start_col, end_row=row, end_column=col - 1)
                    merge_start_col = None
        if merge_start_col is not None:
            ws.merge_cells(start_row=row, start_column=merge_start_col, end_row=row, end_column=end_col)

def apply_column_merges(ws, start_row, end_row, columns_to_merge):
    """Merges specific columns across the given row range."""
    for col in columns_to_merge:
        col_index = column_index_from_string(col)
        ws.merge_cells(start_row=start_row, start_column=col_index, end_row=end_row, end_column=col_index)
        merged_cell = ws.cell(row=start_row, column=col_index)
        merged_cell.value = ws.cell(row=start_row, column=col_index).value

def apply_border_and_alignment(ws, start_row, end_row, start_col, end_col, border_style, alignment_style):
    """Applies border and alignment to the specified cell range."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border_style
            cell.alignment = alignment_style

def process_bit_offsets_and_update(ws_output, bit_offsets, bit_widths, field_names,access_permissions,field_description, start_row, end_row, start_col, end_col, blue_fill, yellow_fill):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    """Processes bit offsets and updates the output sheet."""
    row_index = start_row  # Starting row for inserting data

    for i in range(len(bit_offsets)):
        bit_offset = bit_offsets[i]
        bit_width = bit_widths[i] if i < len(bit_widths) else None
        field_name = field_names[i] if i < len(field_names) else ''
        access_permission = access_permissions[i] if i < len(access_permissions) else ''
        description = field_description[i] if i < len(field_description) else ''  # Fixed variable name


        if bit_offset is None or bit_offset < 0 or bit_offset > 31:
            continue

        start_col_index = end_col - bit_offset
        end_col_index = start_col_index - bit_width + 1 if bit_width else start_col_index

        if bit_offset == 0 and i != 0:
            row_index += 1

        ws_output.merge_cells(start_row=row_index, start_column=end_col_index, end_row=row_index, end_column=start_col_index)
        merged_cell = ws_output.cell(row=row_index, column=end_col_index)
        merged_cell.value = field_name
        merged_cell.fill = blue_fill

      

        if access_permission or description:
            comment_text = f"Bit permission: {access_permission}\nBit Description:\n{description}"
            comment = Comment(comment_text, "Author")
            merged_cell.comment = comment
            comment.width = 500 
            comment.height = 100 
           


    # Color remaining unused cells
    for row_index in range(start_row, end_row + 1):
        for col_index in range(start_col, end_col + 1):
            cell = ws_output.cell(row=row_index, column=col_index)
            if cell.value is not None and cell.value != "":
                cell.fill = blue_fill
            elif cell.fill.start_color.index == '00000000':
                cell.fill = yellow_fill

def align_overall_sheet(ws, bold_headers=True, bold_first_column=True):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                # Check the length of the cell value
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Adjust the column width (slightly more for better readability)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Optionally, set row heights (if needed)
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20  # Customize the height as needed

    print("Sheet alignment, bolding, and formatting applied successfully.")               
            
            

def main(file_path, new_file_path):
     
    module_positions = find_first_and_last_position_per_module(file_path)
    wb = Workbook()
    ws_output = wb.active
    ws_output.title = "Combined Sheet"


    base_address_values = ["0x0010_3000", "0x0010_2000", "0x0018_b000", "0x0018_4000", "0x0018_7000", "0x0018_2000",
                           "0x0018_8000", "0x0018_a000", "0x0018_9000", "0x001d_2000", "0x001c_0000", "0x001c_8000",
                           "0x001d_1000", "0x001d_0000", "0x001c_4000", "0x001c_c000", "0x0004_5000", "0X0005_0000",]
    
    
    baseadd=["0x0000_3000","0x0000_3000","0x00000000","0x0008_b000","0x0008_4000","0x0008_7000","0x0008_2000","0x0008_8000","0x0008_a000","0x0008_9000",
             "0x000d_2000","0x000c_0000","0x000c_8000","0x000d_1000","0x000d_0000","0x000c_4000","0x000c_c000","0x0000_1000","NA"]
    

    all_sheet_names = list(module_positions.keys())

    # Identify the last sheet name
    last_sheet_name = all_sheet_names[-1]

    # Reorder sheets: First sheet, then last sheet, followed by the remaining sheets (excluding the last sheet)
    if len(all_sheet_names) > 1:
        reordered_sheets = [all_sheet_names[0], last_sheet_name] + all_sheet_names[1:-1]
    else:
        reordered_sheets = all_sheet_names
        # Define styles
    blue_fill = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    centered_alignment = Alignment(horizontal='center', vertical='center')
    title_style = Alignment(horizontal='center', vertical='center')

    output_row = 1
    table_counter = 1

    base_address_iter = iter(base_address_values)
    base_iter=iter(baseadd)
    
        # Process each sheet in the input file
    for sheet_name in reordered_sheets:
     print(f"Processing sheet: {sheet_name}")
    
     # Get the module positions for the current sheet
     positions = module_positions[sheet_name]

     # Iterate over modules within the current sheet
     for module_name, position in positions.items():
              
                first_row_index = position['first']
                last_row_index = position['last']

                # Read data from Excel
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            
                output_row+=2
                df = df.iloc[first_row_index-2:last_row_index-1]
                df = df.iloc[:, [2,3,4,-1, -4]]  # Extract specific columns
               
                base_address_value = next(base_address_iter, None)
                baseaddr_value = next(base_iter, None)
                df = update_dataframe(file_path, df, base_address_value,module_name,baseaddr_value)
                           
                df_new = pd.read_excel(file_path, sheet_name=sheet_name)
                df_new=df_new.iloc[first_row_index-2:last_row_index-1]
                # Ensure 'Bit Width', 'Bit Offset', and 'Field Name' columns exist in the DataFrame
                bit_widths = df_new['Bit Width'].dropna().astype(int).tolist() if 'Bit Width' in df_new.columns else []
                bit_offsets = df_new['Bit Offset'].dropna().astype(int).tolist() if 'Bit Offset' in df_new.columns else []
                field_names = df_new['Field Name'].fillna('').tolist() if 'Field Name' in df_new.columns else []
                access_permissions = df_new['Access'].fillna('').tolist() if 'Access' in df_new.columns else []
                field_description=df_new['Field Description'].fillna('').tolist() if 'Field Description'in df_new.columns else[]
                #print(f"Field Descriptions: {field_description}")


                # Insert title and heading
                heading_row = output_row + 1
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=heading_row):
                    for c_idx, cell in enumerate(row, start=1):
                        cell_value = ws_output.cell(row=r_idx, column=c_idx, value=cell)
                        if r_idx == heading_row:
                            cell_value.font = Font(bold=True)
                

                output_row=heading_row
                start_col=5
                end_col=36
                
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

                # Create a table for columns A, B, C, D
                table_name = f"Table{table_counter}"
                create_table(ws_output, output_row, output_row + len(df) , 1, 4, table_name)
                table_counter += 1

                process_bit_offsets_and_update(ws_output, bit_offsets, bit_widths, field_names,access_permissions,field_description, output_row+1, output_row + len(df), start_col, end_col, blue_fill=blue_fill, yellow_fill=yellow_fill)

                insert_module_title(ws_output, module_name, output_row-1, 1, len(df.columns), title_style)

                merge_adjacent_empty_cells(ws_output, output_row, output_row + len(df) , 1, len(df.columns))

                apply_border_and_alignment(ws_output, output_row, output_row + len(df), 1, len(df.columns), border_style, centered_alignment)

                col_index= ["AM", "AN", "AO", "AP", "AQ"]
 
                title_cell = ws_output.cell(row=output_row-1, column=1)
                title_cell.alignment = Alignment(horizontal='left', vertical='top')
                
                output_row += len(df)+1

                apply_column_merges(ws_output,output_row-len(df),output_row-1,col_index)
                                
    align_overall_sheet(ws_output, bold_headers=True, bold_first_column=True) 
    column = 'C'
    rows_to_update = [168,169,181,182,230,231,243,244]
    new_value = "640"


    for row in rows_to_update:
      cell = f'{column}{row}'
      ws_output[cell] = new_value

    cell_to_update = 'AO256'
    cell = ws_output[cell_to_update]
    cell.value ="0x0004_4000"
    cell_to_update = 'AO266'
    cell = ws_output[cell_to_update]
    cell.value ="0X0005_0000"           

    wb.save(new_file_path)
    print(f"Processed file saved as {new_file_path}")


# file path 
input_file_path = r"C:\Users\SivanrajMariappan\Downloads\excel\registers.xlsx"
output_file_path = r"C:\Users\SivanrajMariappan\Downloads\excel\register_output8.xlsx"
main(input_file_path,output_file_path)

